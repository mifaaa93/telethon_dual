from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
import datetime as dt
from typing import List, Dict, Any
from telethon.tl import types


async def create_excel(data: List[Dict[str, Any]], owners: bool = False, include: list[str]= None) -> BytesIO:
    """
    Создаёт Excel-файл в памяти и возвращает BytesIO с установленным именем.
    Если owners=True:
      - добавляет колонки owner_tg_id, owner_username, owner_first_name первыми,
      - группирует по owner_tg_id (None в конце).
    """
    # Сортировка для группировки по владельцу
    if owners:
        data = sorted(
            data,
            key=lambda r: (
                r.get("owner_tg_id") is None,            # None в конец
                r.get("owner_tg_id") or 0,
                (r.get("title") or "").lower(),
            )
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Links Stat"

    base_headers = [
        "Ссылка",
        "Название",
        "Использовано",
        "Одобрено заявок",
        "Всего посещений",
        "Дата создания",
        "Последняя проверка",
    ]
    if owners:
        headers = ["Создал (tg_id)", "Username", "Имя"] + base_headers
    else:
        headers = base_headers

    ws.append(headers)

    # Заголовок жирным
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Локальный часовой пояс
    local_tz = dt.datetime.now().astimezone().tzinfo
    if include:
        data = [row for row in data if row.get("link") in include]
    total = len(data)
    # Заполнение строк
    for row in data:
        date_created = (
            dt.datetime.fromtimestamp(row["date_created"], tz=dt.timezone.utc)
            .astimezone(local_tz)
            .strftime("%Y-%m-%d %H:%M:%S")
            if row.get("date_created") else ""
        )
        last_synced = (
            dt.datetime.fromtimestamp(row["last_synced_at"], tz=dt.timezone.utc)
            .astimezone(local_tz)
            .strftime("%Y-%m-%d %H:%M:%S")
            if row.get("last_synced_at") else ""
        )

        base_cells = [
            row.get("link", ""),
            row.get("title", ""),
            row.get("usage", 0),
            row.get("approved_request_count", 0),
            row.get("visits_total", 0),
            date_created,
            last_synced,
        ]

        if owners:
            owner_id = row.get("owner_tg_id", "")
            owner_username = row.get("owner_username", "")
            owner_first_name = row.get("owner_first_name", "")
            out_row = [owner_id, owner_username, owner_first_name] + base_cells
        else:
            out_row = base_cells

        ws.append(out_row)

    # Автоширина колонок
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                l = len(str(cell.value)) if cell.value is not None else 0
                if l > max_len:
                    max_len = l
            except Exception:
                pass
        ws.column_dimensions[letter].width = max_len + 2

    # В память
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = f"Total_links_stat({total}).xlsx" if owners else f"links_stat_({total}).xlsx"
    return buf


async def create_excel_from_(data: List[types.ChatInviteExported]) -> BytesIO:
    """
    Создаёт Excel-файл в памяти и возвращает BytesIO с установленным именем.
    По списку объектов ChatInviteExported.
    Если owners=True — добавляет колонку admin_id (создатель ссылки) первой.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Links"

    headers = [
        "Ссылка",
        "Название",
        "Использовано",
        "Одобрено заявок",
        "Всего посещений",
        "Дата создания",
    ]
    ws.append(headers)

    # жирная шапка
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # локальный часовой пояс
    local_tz = dt.datetime.now().astimezone().tzinfo

    # строки
    for inv in data:
        link = getattr(inv, "link", "") or ""
        title = getattr(inv, "title", "") or ""
        usage = getattr(inv, "usage", 0) or 0
        approved = getattr(inv, "approved_request_count", 0) or 0
        request_needed = bool(getattr(inv, "request_needed", False))

        visits_total = usage + (approved if request_needed else 0)

        date_obj = getattr(inv, "date", None)
        date_created = (
            dt.datetime.fromtimestamp(int(date_obj.timestamp()), tz=dt.timezone.utc)
              .astimezone(local_tz).strftime("%Y-%m-%d %H:%M:%S")
            if date_obj else ""
        )

        row = [link, title, usage, approved, visits_total, date_created]

        ws.append(row)

    # автоширина колонок
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                l = len(str(cell.value)) if cell.value is not None else 0
                if l > max_len:
                    max_len = l
            except Exception:
                pass
        ws.column_dimensions[letter].width = max_len + 2

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "links_from_exported.xlsx"
    return buf